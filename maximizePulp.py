from math import ceil
import openpyxl
import numpy as np
from pulp import *
# dizionario in cui si assegna uno slot ad un numero
# ogni slot dura da x fino a x + 01:00
slots_half_hour = { # tutti slot da un'ora
    "09:00" : 0, # 09:00 - 09:30
    "09:30" : 1, # 09:30 - 10:00
    "10:00" : 2, # ...
    "10:30" : 3,
    "11:00" : 4,
    "11:30" : 5,
    "14:00" : 6,
    "14:30" : 7,
    "15:00" : 8,
    "15:30" : 9,
    "16:00" : 10,
    "16:30" : 11,
    "17:00" : 12,
    "17:30" : 13,
    "18:00" : 14,
    "18:30" : 15,
    "19:00" : 16,
    "19:30" : 17 # 19:30 - 20:00
}
# stampa un array a 3 dimensioni,hour mi dice la durata dell'allenamento giusto per arrayPosition
def print_3d_array(array,arrayPosition,hour):
    cont = 0
    for s,i in enumerate(arrayPosition):
        if i == hour:
            print (f"giocatore {s+1}")
            print(array[cont,:,:])
            cont += 1
            print('-' * 40)

# stampa la comparazione tra la gli array a 3 dimensioni di allievo e campo (o di 60 o di 90, dipende da cosa si è passato in typeslot) 
# stamperà per ogni studente il numero massimo di allenamenti, la sua matrice di disponibilità e la matrice con gli allenamenti effettivamente assegnati
def print_comparison_3d_array(array1,array2,trainings,arrayPosition,hour):
    cont = 0
    for s,i in enumerate(arrayPosition):
        if i == hour:
            print (f"studente {s+1} può fare al più {trainings[cont]} allenamenti\n")
            for j in range(6):
                print(array1[cont,j,:],array2[cont,j,:])
            cont += 1
            print("\n")
# ci ritorna la variabile decisionale richiesta nella lista a 3 dimensioni delle variabili decisionali
def receive_decision_variables(array,student,day,slot):
    for s,i in enumerate(array):
        if s == student: # ci prendiamo quel particolare studente
            for d,j in enumerate(i):    
                if d == day:    # ci prendiamo quel particolare giorno
                    #print(f"decision variable{j[slot]}")
                    return j[slot]  # ci prendiamo la variabile decisionale di quel particolare slot,del giorno j e dello studente k (x_i,j,k)

# ci ritorna la variabile decisionale richiesta nella lista a 3 dimensioni delle variabili decisionali
def receive_slots_decision_variables(array,lista):
    l = []
    for x in lista:
        day = (x.rsplit(','))[1]
        slot = (x.rsplit(','))[2]

        for giorno in array:
            for ora in giorno:
                if day == ((ora.name).rsplit(",")[0])[1:] and slot == ((ora.name).rsplit(",")[1]):
                    #print( ((ora.name).rsplit(",")[0])[:-1] , ((ora.name).rsplit(",")[1]) )
                    l.append(ora)
    return l

# assegnamo alla matrice allievo di un particolare slot (60 o 90) le disponibilità degli studenti
def assigning_allievo(array,x,i,j,z): 
    if(x == "no" or x == None): # case "no" or None in excel cell
        array[i,j,:] = 0
        return
            
    elif("poi" in x): # case "da x in poi"
        slot = x[3:8]
        array[i,j,slots_half_hour[slot]:] = 1
    
    elif(x == "tutto il pome"): # case "tutto il pome"
        if z == "90":
            
            array[i,j,slots_half_hour["14:00"]:slots_half_hour["18:30"]] = 1
        else : 
            array[i,j,slots_half_hour["14:00"]:slots_half_hour["18:00"]] = 1
            
    elif("giorno" in x): # case "tutto il giorno"
        array[i,j,:] = 1
        
    elif("dopo" in x): # case "da x o dopo y"
        firstHour = x[3:8]
        secondHour = x[16:21]
        array[i,j,slots_half_hour[firstHour]:] = 1
    
    else: #case "da x a y"
        firstHour = x[3:8]
        secondHour = x[11:16]
       
        if secondHour == "12:00":
            array[i,j, slots_half_hour[firstHour] : slots_half_hour["11:30"] + 1] = 1
            
        elif secondHour == "20:00":
            array[i,j, slots_half_hour[firstHour] : slots_half_hour["19:30"] + 1] = 1
            
        else: 
            array[i,j,slots_half_hour[firstHour] : slots_half_hour[secondHour]] = 1
    

def check_dissatisfied(array, nstudents, ntraining, nslots):
    
       for i in range(nstudents):
        sum = 0   
        for j in range(6):
            for k in range(18):
                if array[i,j,k] == 1:
                    
                    sum += 1

        if sum != (ntraining[i] * nslots):
            print(f"studente{i} rimarrà insodisfatto")
            array[i,:,:] = 0
            
        
# per aprire il workbook 
workBook = openpyxl.load_workbook("vincoli2.xlsx")   # Workbook oggetto
third_sheet = workBook.sheetnames[2]
#second_sheet = workBook.sheetnames[1]               # prendiamo il nome del foglio desiderato del file excel
active_worksheet = workBook[third_sheet]             # impostiamo il foglio corrente come attivo
number_students_total = active_worksheet.max_row - 2 # numero totale degli studenti

contentRow = [] # conterrà i valori della tupla man mano che scannerizza le righe di excel

number_students1 = 0    # numero degli studenti che fanno allenamenti da un'ora
number_students130 = 0  # numero degli studenti che fanno allenamenti da un'ora e mezza
exact_position =[] # raccoglie la posizione esatta degli studenti con allenamenti da un'ora e anche da un'ora e mezza

# recuperare i dati dal foglio excel
for i,row in enumerate (active_worksheet.iter_rows(min_row=3, max_col=active_worksheet.max_column, max_row=active_worksheet.max_row)):
    for cell in row:
        contentRow.append(cell.value) # una riga completa

    if contentRow[3] == 90:
        number_students130 += 1
        exact_position.append(9)
    else: # 60
        number_students1 += 1
        exact_position.append(6)

    contentRow.clear()

#print(f"studenti che fanno allenamento da un'ora => {number_students1}",sixty)    
#print(f"studenti che fanno allenamento da un'ora e mezza => {number_students130}",ninety)

# creazione array dove ogni posizione è uno studente e il suo valore è il numero
# massimo di allenamenti che fa in settimana 
number_training1 = np.zeros(number_students1,dtype = int)
number_training130 = np.zeros(number_students130,dtype = int)

# creazione degli array a 3 dimensioni dove 
# ogni posizione è 1 se uno certo studente i è disponibile
# di fare allenamento per quel giorno e per quel slot
# 0 altrimenti
Allievo1 = np.zeros((number_students1,6,18),dtype = int)
Allievo130 = np.zeros((number_students130,6,18),dtype = int)
# array a 3 dimensioni che verranno popolate dalla funzione obiettivo
Campo1 = np.zeros((number_students_total,6,18),dtype = int)
Campo130 = np.zeros((number_students_total,6,18),dtype = int)
#Slots = np.zeros((6,9), dtype = int)

cont1 = 0
cont130 = 0
switch = True

# assegnamo agli array 3d allievo1 e 130 le diponibilità da parte degli allievi
for i,row in enumerate (active_worksheet.iter_rows(min_row=3, max_col=active_worksheet.max_column, max_row=active_worksheet.max_row)):
    # aggiungiamo a contentRow tutti i valori della tupla che viene estratta da iter_rows
    for cell in row:
        contentRow.append(cell.value)

    # prendere il numero di allenemanti massimi di ogni studente e se fa allenamento da un'ora opuure un'ora e mezza
    if "1" in contentRow[2] or "2" in contentRow[2] or "3" in contentRow[2] or "4" in contentRow[2]:
        trainings = contentRow[2]

        if contentRow[3] == 90: number_training130[cont130] = int(trainings[0:1])     
        else: number_training1[cont1] = int(trainings[0:1])

    # lettura degli slot (un sottoinsieme di una riga!)
    for j,x in enumerate(contentRow[4:10]):
        if contentRow[3] == 90:
            assigning_allievo(Allievo130,x,cont130,j,"90")
            switch = True
        else:
            assigning_allievo(Allievo1,x,cont1,j,"60")
            switch = False

    if(switch):
        cont130 +=1
    else :
        cont1 += 1

    contentRow.clear()

print_3d_array(Allievo1,exact_position,6)
print_3d_array(Allievo130,exact_position,9)
print(number_training1)
print(number_training130)

# creazioni dei problemi di programmazione intera
problem1 = LpProblem("Maximize_student_1", sense=LpMaximize)
problem130 = LpProblem("Maximize_student_130", sense=LpMaximize)
i = 0
j = 0 
# variabili decisionali, x per quelli da un'ora e y per quelli da un'ora e mezza
students_variables1 =[
    [   
    [LpVariable(name=f"x_{i+1},{j+1},{k+1}", cat=LpBinary) for k in range(18)]
        for j in range(6)
    ] for i in range(number_students1)
]
students_variables130 =[
    [   
    [LpVariable(name=f"y_{i+1},{j+1},{k+1}", cat=LpBinary) for k in range(18)]
        for j in range(6)
    ] for i in range(number_students130)
]
# funzioni obiettivo da massimizzare
problem1 += lpDot(1,students_variables1)
problem130 += lpDot(1,students_variables130)

#---------------------------vincoli------------------------
# 2) vincolo // ogni studente si allena al massimo una volta al giorno
# per farlo dobbiamo controllare l'array 3d allievo e vedere quale studente per un certo giorno per un certo slot è libero (1, 0 altrimenti)
# una volta capito se è libero dobbiamo prendere la variabile decisionale corrispondente (di quello studente per quel certo giorno per quel certo slot) 
# e mettere tutte le variabili decisionali di uno studente di un giorno (o più) in una espressione lineare e controllare che sia <=1
# 1) vincolo // aggiunto se no uno studente non ha disponibilità per un certo giorno per uno certo slot la variabile decisionale corrispondente è uguale a 0
# i = studente, j = giorno, k = slot
students_one = []
students_zero = []
for s,i in enumerate(Allievo1):
    #print(f"studente = {s+ 1}")
    for d,j in enumerate(i):        
        for o,k in enumerate(j):
            if k == 0:
                students_zero.append(receive_decision_variables(students_variables1,s,d,o))
            else: # k == 1
                students_one.append(receive_decision_variables(students_variables1,s,d,o))
                
        #print(students_one)
        if not(len(students_zero) == 0):
            problem1 += lpSum(1 * students_zero) == 0
        
        if not(len(students_one) == 0):

            if len(students_one) >= 2:
                
                for z in range(0,len(students_one),2):
                    #if z == (len(students_one) - 1): break
                    problem1 += lpDot(1, students_one[z]) == lpDot(1, students_one[z + 1])
                    
            problem1 += lpDot(1,students_one) <= 2 #solo un allenamento al giorno (2 slot) 
        students_one.clear()
        students_zero.clear()
# vincolo 3 // per ogni slot ci devono essere al massimo 4 allievi
for day in range(6):
    for slot in range(18):
        for student in range(number_students1):
            if(Allievo1[student,day,slot] == 1):
                students_one.append(receive_decision_variables(students_variables1,student,day,slot))
        if not(len(students_one) == 0):
            problem1 += lpSum(1 * students_one) <= 4
            students_one.clear()
    
# vincolo 4 // uno studente si allena al massimo il numero di allenamenti da lui dichiarato
for s,i in enumerate(Allievo1):
    #print(f"studente = {s+ 1}")
    for d,j in enumerate(i):        
        for o,k in enumerate(j):
            if k == 1:
                students_one.append(receive_decision_variables(students_variables1,s,d,o)) 

    problem1 += lpSum(1 * students_one) <= (number_training1[s] * 2)
    students_one.clear()

# stessa cosa ma per gli studenti da un'ora e mezza
students_one = []
students_zero = []
for s,i in enumerate(Allievo130):
    #print(f"studente = {s+ 1}")
    for d,j in enumerate(i):        
        for o,k in enumerate(j):
            if k == 0:
                students_zero.append(receive_decision_variables(students_variables130,s,d,o))
            else: # k == 1
                students_one.append(receive_decision_variables(students_variables130,s,d,o))

        #print(students_one)
        if not(len(students_zero) == 0):
            problem130 += lpSum(1 * students_zero) == 0
        
        if not(len(students_one) == 0):

            if len(students_one) >= 3:
                
                for z in range(0,len(students_one),3):
                    if z == (len(students_one) - 1): break
                    #print(students_one)
                    problem130 += lpDot(1, students_one[z]) == lpDot(1, students_one[z + 1])
                    problem130 += lpDot(1, students_one[z + 1]) == lpDot(1, students_one[z + 2])
                    
            problem130 += lpDot(1,students_one) <= 3 #solo un allenamento al giorno (3 slot)
        
        students_one.clear()
        students_zero.clear()

for day in range(6):
    for slot in range(18):
        for student in range(number_students130):
            if(Allievo130[student,day,slot] == 1):
                students_one.append(receive_decision_variables(students_variables130,student,day,slot))
        if not(len(students_one) == 0):
            problem130 += lpSum(1 * students_one) <= 4
            students_one.clear()
    

for s,i in enumerate(Allievo130):
    #print(f"studente = {s+ 1}")
    for d,j in enumerate(i):        
        for o,k in enumerate(j):
            if k == 1:
                students_one.append(receive_decision_variables(students_variables130,s,d,o)) 

    problem130 += lpSum(1 * students_one) <= (number_training130[s] * 3)
    students_one.clear()

# risoluzione dei problemi
status1 = problem1.solve(PULP_CBC_CMD(msg = False))
status130 = problem130.solve(PULP_CBC_CMD(msg = False))

# riempire la matrice campo (il risultato con le variabili decisionali scelte)
# 60
for v in problem1.variables():
    
    if v.varValue == 1.0:
        #print(v.name)
        temp = (v.name).rsplit(",")
        Campo1[int((temp[0])[2:]) - 1,int(temp[1]) - 1,int(temp[2]) - 1] = 1
# 90
for v in problem130.variables():
    if v.varValue == 1.0:
        temp = (v.name).rsplit(",")
        Campo130[int((temp[0])[2:]) - 1,int(temp[1]) - 1,int(temp[2]) - 1] = 1
        
#print_3d_array(Campo130,ninety)
#print_3d_array(Campo1,sixty)
#print(problem1)
print(problem130)
s1 = ceil(number_students1 / 4)
s15 = ceil(number_students130 / 4)


check_dissatisfied(Campo1, number_students1, number_training1, 2) 
check_dissatisfied(Campo130, number_students130, number_training130, 3)

print_comparison_3d_array(Allievo1,Campo1,number_training1,exact_position,6)
print("|" * 80,'\n')
print_comparison_3d_array(Allievo130,Campo130,number_training130,exact_position,9)


print ("valore ottimo per gli studenti da un'ora e mezza = ", problem130.objective.value())
print ("valore ottimo per gli studenti da un'ora= ", problem1.objective.value())
print(s1,s15)



decision_variables_students_final_1 = []
decision_variables_students_final_130 = []

for i in range(number_students_total):
    for j in range(6):
        for k in range(18):
            if Campo1[i,j,k] == 1: 
                decision_variables_students_final_1.append(receive_decision_variables(students_variables1,i,j,k))

for i in range(number_students_total):
    for j in range(6):
        for k in range(18):
            if Campo130[i,j,k] == 1: 
                decision_variables_students_final_130.append(receive_decision_variables(students_variables130,i,j,k))

#print(len(decision_variables_students_final_1),decision_variables_students_final_1)
#print(len(decision_variables_students_final_130),decision_variables_students_final_130)
# in questi for vado a modificare sixty e ninety
# perchè è possibile che ci sia qualche studente non soddisfatto (es: dichiarato due allenamenti ma assegnato solo uno) e quindi aggiornamo
# nella posizione esatta il fatto che non ci sia più
zeros = 0
stu6 = 0
stu9 = 0
for s,i in enumerate(exact_position):
    if  i == 6:
        zeros = 0
        for j in range(6):
            for k in range(18): 
                if Campo1[stu6,j,k] == 0:
                    zeros += 1
                else: break
        if zeros == 108:
            exact_position[s] = 0
        stu6 += 1
    else:
        zeros = 0
        for j in range(6):
            for k in range(18): 
                if Campo130[stu9,j,k] == 0:
                    zeros += 1
                else: break
        if zeros == 108:
            exact_position[s] = 0
        stu9 += 1


cont1 = 0
cont130 = 0
end1 = []
end130 = []

for s,i in enumerate(exact_position):
    if i == 6 :
        if cont1 == len(decision_variables_students_final_1):
                break
        temp = str(decision_variables_students_final_1[cont1]).rsplit(",")
        stud = (temp[0])[2:]
        
        st = "x_" + str(f"{s + 1},{temp[1]},{temp[2]}")
        end1.append(st)
        cont1 += 1
        
        while stud == ((str(decision_variables_students_final_1[cont1]).rsplit(","))[0])[2:] :
            
            temp = str(decision_variables_students_final_1[cont1]).rsplit(",")
            st = "x_" + str(f"{s + 1},{temp[1]},{temp[2]}")
            end1.append(st)
            cont1 += 1
            if cont1 == len(decision_variables_students_final_1):
                break
    else:

        if cont130 == len(decision_variables_students_final_130):
                break
        temp = str(decision_variables_students_final_130[cont130]).rsplit(",")
        stud = (temp[0])[2:]
        
        st = "x_" + str(f"{s + 1},{temp[1]},{temp[2]}")
        end130.append(st)
        cont130 += 1
        
        while stud == ((str(decision_variables_students_final_130[cont130]).rsplit(","))[0])[2:] :
            
            temp = str(decision_variables_students_final_130[cont130]).rsplit(",")
            st = "x_" + str(f"{s + 1},{temp[1]},{temp[2]}")
            end130.append(st)
            cont130 += 1
            if cont130 == len(decision_variables_students_final_130):
                break  

print(end1)
print(end130)
#print_3d_array(Campo_1final,sixty)
#print_3d_array(Campo_130final,ninety)
'''
print(final_xvariables)
print(f"lunghezza x {len(final_xvariables)}"
print(final_yvariables)
print(f"lunghezza y {len(final_yvariables)}")
'''

stu_content = []
day_content = []
final = []

save_stu = ((end1[0].rsplit(","))[0])[2:]
save_day = ((end1[0].rsplit(","))[1])

for i in end1 :
    temp = i.rsplit(",")
    stu = (temp[0])[2:]
    day = temp[1]
    if (stu == save_stu):
        if (day == save_day):
            stu_content.append(i)
        else:
            save_day = day
            day_content.append(stu_content)
            stu_content = []
            stu_content.append(i)

    else:
        day_content.append(stu_content)
        final.append(day_content)
        #print(final)
        save_day = day
        save_stu = stu
        stu_content = []
        day_content = [] 
        stu_content.append(i)

day_content.append(stu_content)
final.append(day_content)

x1 = final

final = []
day_content = []
stu_content = []

save_stu = ((end130[0].rsplit(","))[0])[2:]
save_day = ((end130[0].rsplit(","))[1])

for i in end130 :
    temp = i.rsplit(",")
    stu = (temp[0])[2:]
    day = temp[1]
    if (stu == save_stu):
        if (day == save_day):
            stu_content.append(i)
        else:
            save_day = day
            day_content.append(stu_content)
            stu_content = []
            stu_content.append(i)

    else:
        day_content.append(stu_content)
        final.append(day_content)
        
        #print(final)
        save_day = day
        save_stu = stu
        stu_content = []
        day_content = []
        stu_content.append(i)

day_content.append(stu_content)
final.append(day_content)

x130 = final

print(x1,"\n",x130,"\n")

final_problem = LpProblem("resfinale", sense=LpMinimize)

slots_variable = [
    [LpVariable(name=f"y{j+1},{k+1}", cat=LpBinary) for k in range(18)]
        for j in range (6)
    ]
x_variables = [LpVariable(name= f"x_{i+1}", cat=LpBinary) for i in range (150)]

# funzione obiettivo : minimizzare il numero di slot ovvero di ore utlizzate
final_problem += lpDot(1,slots_variable)
print(final_problem)

minimumslot =[s1,s15]
print(minimumslot)
if (54 <= s1 + s15):
    print(f"il numero totali è minore della somma degli slot minimi")
    exit()

stop = False
halt = False
#Vincoli
cont = 0
sus = []
already_used = []
for i in x1:
    #print(i)
    for check1 in i:
        print(f"\ncontrolliamo {check1}")
        first = receive_slots_decision_variables(slots_variable,check1)
        print(first)
        
        for item in first:
            for item2 in already_used:
                if str(item) == str(item2):
                    #print(item, item2)
                    #print(f"ci sono già {item}")
                    halt = True
                    break
            if(halt) : break
        if (halt):
            halt = False
            continue
        else:
            already_used.extend(first)
        #print(f"porco dio {already_used}")
        #final_problem += lpDot(1,first[0]) == lpDot(1,first[1])
        sus.append(x_variables[cont])   
        for j in check1:
            if (stop):
                stop = False
                break
            else:
                day = (j.rsplit(","))[1]
                slot = (j.rsplit(","))[2]
    
            for t in x130: # uno studente da tre slot 
                #print(t)
                for check130 in t: # per un certo giorno di quel studente  
                    if (stop) : break
                    for z in check130: # gli slot di quel giorno di quel studente
                        giorno = (z.rsplit(","))[1]
                        ora = (z.rsplit(","))[2]
                        if day == giorno and slot == ora:
                            print(check130)
                            final_problem += lpDot(1,first) == lpSum(2 * x_variables[cont])
                            cont += 1
                            second = receive_slots_decision_variables(slots_variable,check130)
                            sus.append(x_variables[cont])
                            final_problem += lpDot(1,second) == lpSum(3 * x_variables[cont])
                            #final_problem += lpDot(1,second[0]) == lpDot(1,second[1])
                            #final_problem += lpDot(1,second[1]) == lpDot(1,second[2])
                            cont += 1
                            print(second)
                            stop = True
                            break
                    

        if(len(second) == 0):
            print("no match")
            final_problem += lpDot(1,first) == 2
        else:
            #final_problem += lpDot (1,sus[0]) <= lpDot (1,sus[1])
            
            final_problem += lpDot(1,sus) >= 1
            '''
            for item in sus:
                final_problem += lpDot(1,item) == 1
            '''
              
            sus.clear()
            #print('lul',second)
            second.clear()
            break               
                    
print(final_problem)
solve = final_problem.solve(PULP_CBC_CMD(msg = False))

for v in  final_problem.variables():
    print(v.name, "=", v.varValue)

print ("valore ottimo degli slot minimi ", final_problem.objective.value())

res = np.zeros((6,18),dtype = int)

x = 0
y = 0

for v in final_problem.variables():
    if "y" in v.name and v.varValue == 1:
        temp = (v.name).rsplit(",")
        #print(temp[0]+","+temp[1])
        res[int((temp[0])[1:]) - 1][int(temp[1]) - 1] = 1
    
print(res)